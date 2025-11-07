# -*- coding: utf-8 -*-
"""
Created on Tue Oct 14 15:00:40 2025

@author: pablo.SAIDI
"""

import openpyxl
from tkinter import Tk, filedialog
import os
import sys

# === Excel file selection window ===
root = Tk()
root.withdraw()
root.call('wm', 'attributes', '.', '-topmost', True)  # keep window on top

file_path = filedialog.askopenfilename(
    title="Select the Excel file to process",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)
root.destroy()

if not file_path:
    print("❌ No file selected. Script terminated.")
    sys.exit()

# === Open the Excel file ===
wb = openpyxl.load_workbook(file_path)
sheet = wb.active  # or wb["Your_Sheet_Name"]

# === Automatically read animal weights from B3:C7 ===
animal_weights = {}
for row in range(3, 8):  # from B3 to C7
    box = sheet[f"B{row}"].value
    weight = sheet[f"C{row}"].value
    if box is not None and weight is not None:
        try:
            animal_weights[int(box)] = float(weight)
        except:
            pass

print("📦 Detected weights:", animal_weights)

if not animal_weights:
    raise ValueError("❌ No weights detected in cells B3:C7.")

# === Automatically find "Box" and "VO2(1)" columns from row 9 ===
col_box = None
col_vo2 = None

for cell in sheet[9]:
    if cell.value and str(cell.value).strip().lower() == "box":
        col_box = cell.column
    elif cell.value and "vo2" in str(cell.value).lower():
        col_vo2 = cell.column

if col_box is None or col_vo2 is None:
    raise ValueError("❌ Could not find 'Box' and 'VO2(1)' columns from row 9.")

# === Write the header and unit of the new column (Q9 and Q10) ===
sheet["Q9"] = "Energy expenditure"
sheet["Q10"] = "[kcal/h]"

# === Calculate and write values starting from row 11 ===
row = 11
while sheet.cell(row=row, column=col_vo2).value not in (None, ""):
    try:
        vo2 = float(sheet.cell(row=row, column=col_vo2).value)
        box = int(sheet.cell(row=row, column=col_box).value)
        weight = animal_weights.get(box, 0)
        expenditure = vo2 * weight * 0.000005
        sheet[f"Q{row}"] = expenditure
    except Exception:
        sheet[f"Q{row}"] = None
    row += 1

# === Save the result in the same folder ===
folder = os.path.dirname(file_path)
file_name = os.path.basename(file_path)
output_name = file_name.replace(".xlsx", "_results.xlsx")
output_path = os.path.join(folder, output_name)

wb.save(output_path)

print("\n✅ Column 'Energy expenditure [kcal/h]' added from Q9–Q10!")
print(f"📁 File saved at: {output_path}")
