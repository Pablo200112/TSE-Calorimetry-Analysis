# -*- coding: utf-8 -*-
"""
Created on Tue Oct 14 10:32:40 2025
@author: pablo.SAIDI
"""

import pandas as pd
from openpyxl import load_workbook
from tkinter import Tk, filedialog
import os

# --- File selection ---
Tk().withdraw()

print("📂 Select the main file (e.g., PS 2025 01 arvis M.xlsx)")
file1 = filedialog.askopenfilename(
    title="Select the main file",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

print("📂 Select the file to add (e.g., PS 2025 01 arvis M bis.xlsx)")
file2 = filedialog.askopenfilename(
    title="Select the file to add",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

if not file1 or not file2:
    raise SystemExit("❌ Selection cancelled. Restart the script and choose both Excel files.")

# --- Read both files starting from row 9 ---
df1 = pd.read_excel(file1, skiprows=8)
df2 = pd.read_excel(file2, skiprows=8)

df1.columns = df1.columns.str.strip()
df2.columns = df2.columns.str.strip()

# --- Detect the column containing the animal name or code ---
animal_col = None
for col in df1.columns:
    if 'animal' in col.lower():
        animal_col = col
        break

if not animal_col:
    print("❌ Could not find the column containing the animal name.")
    print("Available columns:", df1.columns.tolist())
    raise SystemExit

# --- Open the main file with openpyxl ---
wb = load_workbook(file1)
ws = wb.active

print("\n🔍 Inserting data from the second file into the correct sections...")

# Get the order of animals in the main file
animal_order = df1[animal_col].dropna().unique().tolist()

# Insert starting from the end to avoid shifting already processed rows
for animal in reversed(animal_order):
    new_rows = df2[df2[animal_col] == animal]
    if new_rows.empty:
        continue

    # Find the last row for this animal in the main file
    animal_rows = df1[df1[animal_col] == animal]
    if animal_rows.empty:
        continue

    last_row_index = animal_rows.index[-1] + 10  # +9 for skipped rows +1 because Excel is 1-based

    # Insert new rows just after
    ws.insert_rows(last_row_index + 1, amount=len(new_rows))

    for i, (_, row) in enumerate(new_rows.iterrows(), start=0):
        for j, val in enumerate(row, start=1):
            ws.cell(row=last_row_index + 1 + i, column=j, value=val)

    print(f"✅ Data for {animal} inserted after row {last_row_index}")

# --- Save the result ---
# The final file will take the name of the first file + " - final compiled.xlsx"
base_name = os.path.splitext(os.path.basename(file1))[0]
desktop_path = r"D:\pablo.SAIDI\Desktop"
output_file = os.path.join(desktop_path, f"{base_name} - final merged.xlsx")

wb.save(output_file)

print("\n🎉 Merge completed successfully!")
print(f"💾 The final file has been saved here:\n{output_file}")
print("✅ New data has been inserted right after the correct animals, keeping the original structure intact.")
